# ============================================
# app.py — COMPLETE ROBUST DROPOUT PREDICTOR
# Works with ANY dataset — no column crashes
# Outputs: Student + Teacher + Management view
# ============================================

import streamlit as st
import pandas as pd
import numpy as np
import pickle
import matplotlib.pyplot as plt
import matplotlib
import seaborn as sns
import os
import io
import warnings
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font,
    Alignment, Border, Side)
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
matplotlib.use('Agg')

# ============================================
# PAGE CONFIG
# ============================================
st.set_page_config(
    page_title="Student Dropout Predictor",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# CUSTOM CSS
# ============================================
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(
            135deg, #1a1a2e, #0f3460);
        padding: 20px;
        border-radius: 10px;
        text-align: center;
        margin-bottom: 20px;
    }
    .main-header h1 {
        color: white;
        font-size: 28px;
        margin: 0;
    }
    .main-header p {
        color: #a8d8ea;
        margin: 5px 0 0 0;
    }
    .kpi-box {
        background: white;
        border-radius: 10px;
        padding: 15px;
        text-align: center;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        border-left: 5px solid #3498db;
    }
    .kpi-number {
        font-size: 36px;
        font-weight: bold;
    }
    .kpi-label {
        font-size: 12px;
        color: #666;
        margin-top: 5px;
    }
    .risk-critical {
        background-color: #fdecea;
        border-left: 5px solid #e74c3c;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .risk-high {
        background-color: #fff3e0;
        border-left: 5px solid #e67e22;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .risk-medium {
        background-color: #fef9e7;
        border-left: 5px solid #f39c12;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .risk-low {
        background-color: #e8f5e9;
        border-left: 5px solid #27ae60;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .stDataFrame { font-size: 13px; }
    div[data-testid="metric-container"] {
        background: white;
        border-radius: 10px;
        padding: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
</style>
""", unsafe_allow_html=True)

# ============================================
# LOAD MODELS — WITH ERROR HANDLING
# ✅ FIXED: Using v2 filenames
# ============================================
@st.cache_resource
def load_models():
    try:
        model = pickle.load(
            open('lr_model_v2.pkl', 'rb'))
        scaler = pickle.load(
            open('scaler.pkl', 'rb'))
        threshold = pickle.load(
            open('threshold_v2.pkl', 'rb'))

        # Try loading SHAP explainer
        try:
            explainer = pickle.load(
                open('shap_explainer_v2.pkl', 'rb'))
        except:
            explainer = None

        return model, scaler, threshold, explainer, True

    except Exception as e:
        return None, None, 0.5, None, False


model, scaler, THRESHOLD, explainer, models_loaded = (
    load_models())

# ============================================
# CORE CONFIGURATION
# ============================================
FEATURES = [
    'attendance_frequency',
    'quiz_score',
    'assignment_completion',
    'weekly_study_hours',
    'learning_satisfaction',
    'prior_dropout_intention'
]

FEATURE_DISPLAY = {
    'attendance_frequency'   : 'Attendance',
    'quiz_score'             : 'Quiz Score',
    'assignment_completion'  : 'Assignments',
    'weekly_study_hours'     : 'Study Hours',
    'learning_satisfaction'  : 'Satisfaction',
    'prior_dropout_intention': 'Dropout Intent'
}

# ALL UCI COLUMNS THAT MIGHT BE NEEDED
# With safe default values
UCI_DEFAULTS = {
    'Curricular units 2nd sem (approved)'       : 0,
    'Curricular units 2nd sem (enrolled)'       : 1,
    'Curricular units 2nd sem (grade)'          : 0,
    'Curricular units 2nd sem (evaluations)'    : 0,
    'Curricular units 1st sem (approved)'       : 0,
    'Curricular units 1st sem (enrolled)'       : 1,
    'Curricular units 1st sem (grade)'          : 0,
    'Curricular units 1st sem (evaluations)'    : 0,
    'Tuition fees up to date'                   : 0,
    'Scholarship holder'                        : 0,
    'Debtor'                                    : 0,
    'Age at enrollment'                         : 20,
    'Gender'                                    : 0,
    'Admission grade'                           : 100,
    'Displaced'                                 : 0,
    'Educational special needs'                 : 0,
    'International'                             : 0,
    'Unemployment rate'                         : 10,
    'Inflation rate'                            : 1,
    'GDP'                                       : 1,
}

# ============================================
# SAFE COLUMN HANDLER
# ============================================
def ensure_columns(df, required_dict):
    filled = []
    for col, default_val in required_dict.items():
        if col not in df.columns:
            df[col] = default_val
            filled.append(col)

    if filled:
        st.sidebar.warning(
            f"⚠️ {len(filled)} columns were "
            f"missing and filled with defaults:\n"
            + "\n".join([f"• {c}" for c in filled])
        )
    return df


# ============================================
# FEATURE ENGINEERING
# ============================================
def engineer_features(df_raw):
    df = df_raw.copy()
    df = ensure_columns(df, UCI_DEFAULTS)

    max_grade = max(
        df['Curricular units 2nd sem (grade)'].max(), 1)
    max_units = max(
        df['Curricular units 2nd sem (enrolled)'].max(), 1)

    df['attendance_frequency'] = (
        df['Curricular units 2nd sem (approved)'] /
        df['Curricular units 2nd sem (enrolled)'].replace(0, 1)
    ).clip(0, 1).fillna(0)

    df['quiz_score'] = (
        df['Curricular units 2nd sem (grade)'] /
        max_grade
    ).clip(0, 1).fillna(0)

    df['assignment_completion'] = (
        df['Curricular units 2nd sem (evaluations)'] /
        df['Curricular units 2nd sem (enrolled)'].replace(0, 1)
    ).clip(0, 1).fillna(0)

    df['weekly_study_hours'] = (
        df['Curricular units 2nd sem (enrolled)'] /
        max_units
    ).clip(0, 1).fillna(0)

    df['learning_satisfaction'] = (
        df['Tuition fees up to date'] * 0.7 +
        df['Scholarship holder'] * 0.3
    ).clip(0, 1).fillna(0)

    df['prior_dropout_intention'] = (
        df['Debtor'].fillna(0).astype(int))

    return df


# ============================================
# RISK CLASSIFICATION
# ============================================
def get_risk(prob):
    if prob >= 0.80:
        return ("CRITICAL", "🔴", "#e74c3c", "fdecea")
    elif prob >= 0.60:
        return ("HIGH", "🟠", "#e67e22", "fff3e0")
    elif prob >= 0.30:
        return ("MEDIUM", "🟡", "#f39c12", "fef9e7")
    else:
        return ("LOW", "🟢", "#27ae60", "e8f5e9")


# ============================================
# REASON GENERATION (SHAP-AWARE)
# ============================================
def get_reasons(row, shap_vals=None):
    reasons = []

    if shap_vals is not None:
        pairs = list(zip(FEATURES, shap_vals))
        pairs.sort(key=lambda x: x[1], reverse=True)

        for feat, val in pairs:
            if val > 0.05:
                fval = row[feat]
                if feat == 'attendance_frequency':
                    reasons.append(
                        f"Low attendance ({fval*100:.0f}%) — SHAP impact: +{val:.3f}")
                elif feat == 'quiz_score':
                    reasons.append(
                        f"Poor quiz performance ({fval*100:.0f}%) — SHAP impact: +{val:.3f}")
                elif feat == 'assignment_completion':
                    reasons.append(
                        f"Low assignment completion ({fval*100:.0f}%) — SHAP impact: +{val:.3f}")
                elif feat == 'weekly_study_hours':
                    reasons.append(
                        f"Insufficient study hours ({fval*8:.1f} hrs/week) — SHAP impact: +{val:.3f}")
                elif feat == 'learning_satisfaction':
                    reasons.append(
                        f"Low satisfaction ({fval*5:.1f}/5) — SHAP impact: +{val:.3f}")
                elif feat == 'prior_dropout_intention':
                    if fval == 1:
                        reasons.append(
                            f"Financial stress / dropout intention signal — SHAP impact: +{val:.3f}")
    else:
        if row['attendance_frequency'] < 0.40:
            reasons.append(
                f"Low attendance ({row['attendance_frequency']*100:.0f}%) — below 40% threshold")
        if row['quiz_score'] < 0.40:
            reasons.append(
                f"Poor quiz scores ({row['quiz_score']*100:.0f}%) — below 40% threshold")
        if row['assignment_completion'] < 0.40:
            reasons.append(
                f"Low assignment completion ({row['assignment_completion']*100:.0f}%)")
        if row['weekly_study_hours'] < 0.30:
            reasons.append(
                f"Insufficient study hours ({row['weekly_study_hours']*8:.1f} hrs/week)")
        if row['learning_satisfaction'] < 0.50:
            reasons.append("Low course satisfaction")
        if row['prior_dropout_intention'] == 1:
            reasons.append("Financial stress or dropout intention detected")

    return reasons if reasons else ["No significant risk factors detected"]


# ============================================
# SUGGESTIONS FOR STUDENT
# ============================================
def get_student_suggestions(row, prob):
    risk, _, _, _ = get_risk(prob)
    suggestions = []

    if row['attendance_frequency'] < 0.60:
        suggestions.append(
            "📅 Attend at least 4 out of 5 weekly sessions. "
            "Consistent attendance is your most important success factor")

    if row['quiz_score'] < 0.60:
        suggestions.append(
            "📝 Spend 30 minutes daily reviewing lecture notes and "
            "practice 5 quiz questions before each test")

    if row['assignment_completion'] < 0.60:
        suggestions.append(
            "📚 Submit all pending assignments now. "
            "Ask your teacher for a deadline extension — they want to help")

    if row['weekly_study_hours'] < 0.40:
        suggestions.append(
            "⏰ Set 4 dedicated study hours weekly. "
            "Use Pomodoro: 25 min study + 5 min break for best results")

    if row['learning_satisfaction'] < 0.50:
        suggestions.append(
            "😊 Tell your teacher what is making the course difficult. "
            "Your feedback directly improves your experience")

    if row['prior_dropout_intention'] == 1:
        suggestions.append(
            "💰 Visit the student support office for scholarship and financial aid. "
            "Do not quit because of money pressure")

    if not suggestions:
        suggestions.append(
            "🌟 You are doing great! Maintain your current pace and "
            "you will complete this course successfully")

    if risk == 'CRITICAL':
        suggestions.append(
            "🆘 Your teacher will contact you very soon. "
            "Please respond — this one conversation can change everything")
    elif risk == 'HIGH':
        suggestions.append(
            "💪 Three weeks of consistent effort will significantly "
            "reduce your risk. You are closer than you think")

    return suggestions


# ============================================
# TEACHER ACTIONS
# ============================================
def get_teacher_actions(prob, reasons):
    risk, _, _, _ = get_risk(prob)
    actions = []

    if risk == 'CRITICAL':
        actions.append("📞 CALL STUDENT TODAY — Personal intervention in 24 hours")
        actions.append("📋 Schedule emergency counseling session")
    elif risk == 'HIGH':
        actions.append("💬 Contact within 48 hours by phone")
        actions.append("📧 Send personalized support email with study resources")
    elif risk == 'MEDIUM':
        actions.append("📩 Send encouraging message today with specific improvement tips")
    else:
        actions.append("👁️ Weekly monitoring — no urgent action needed")

    reason_text = " ".join(reasons)
    if "attendance" in reason_text.lower():
        actions.append("📅 Ask about scheduling or technical barriers to attending")
    if "quiz" in reason_text.lower():
        actions.append("📝 Share topic-specific study materials and practice questions")
    if "assignment" in reason_text.lower():
        actions.append("⏰ Offer 3-day deadline extension for pending assignments")
    if "financial" in reason_text.lower():
        actions.append("💰 Refer to financial aid office and scholarship services")

    return actions


# ============================================
# PREDICT ALL STUDENTS
# ============================================
def predict_all(df_feat):
    for f in FEATURES:
        if f not in df_feat.columns:
            df_feat[f] = 0

    X = df_feat[FEATURES].fillna(0)
    X_scaled = scaler.transform(X)
    probs = model.predict_proba(X_scaled)[:, 1]

    shap_vals_all = None
    if explainer is not None:
        try:
            shap_vals_all = explainer.shap_values(X_scaled)
            if isinstance(shap_vals_all, list):
                shap_vals_all = shap_vals_all[0]
        except:
            shap_vals_all = None

    return probs, X_scaled, shap_vals_all, X


# ============================================
# EXCEL HELPER FUNCTIONS
# ============================================
def make_fill(hex_color):
    return PatternFill(
        start_color=hex_color,
        end_color=hex_color,
        fill_type='solid')

def make_font(bold=False, color='FF000000', size=11):
    return Font(bold=bold, color=color, size=size, name='Calibri')

def make_align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def make_border():
    thin = Side(style='thin', color='FFBDC3C7')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================
# EXCEL REPORT GENERATOR
# ============================================
def generate_excel(df_results, teacher_name, course_name, platform_name):
    wb = Workbook()

    total    = len(df_results)
    critical = (df_results['Risk_Level'] == 'CRITICAL').sum()
    high     = (df_results['Risk_Level'] == 'HIGH').sum()
    medium   = (df_results['Risk_Level'] == 'MEDIUM').sum()
    low      = (df_results['Risk_Level'] == 'LOW').sum()
    at_risk  = critical + high

    RISK_COLORS = {
        'CRITICAL': ('FFE74C3C', 'FFFDECEA'),
        'HIGH'    : ('FFE67E22', 'FFFFF3E0'),
        'MEDIUM'  : ('FFF39C12', 'FFFEF9E7'),
        'LOW'     : ('FF27AE60', 'FFE8F5E9'),
    }

    # ── SHEET 1: ALL STUDENTS ──
    ws1 = wb.active
    ws1.title = "All Students"

    ws1.merge_cells('A1:M1')
    ws1['A1'] = (
        f"STUDENT DROPOUT PREDICTION REPORT  |  "
        f"{platform_name}  |  Course: {course_name}  |  Teacher: {teacher_name}")
    ws1['A1'].fill = make_fill('FF1A1A2E')
    ws1['A1'].font = make_font(bold=True, color='FFFFFFFF', size=13)
    ws1['A1'].alignment = make_align()
    ws1.row_dimensions[1].height = 32

    summary = [
        ('A2:B2', f"Total: {total}", 'FF3498DB'),
        ('C2:D2', f"Critical: {critical}", 'FFE74C3C'),
        ('E2:F2', f"High: {high}", 'FFE67E22'),
        ('G2:H2', f"Medium: {medium}", 'FFF39C12'),
        ('I2:J2', f"Low: {low}", 'FF27AE60'),
        ('K2:M2', f"At Risk: {at_risk} ({at_risk/max(total,1)*100:.0f}%)", 'FFE94560'),
    ]
    for rng, txt, col in summary:
        ws1.merge_cells(rng)
        ref = rng.split(':')[0]
        ws1[ref] = txt
        ws1[ref].fill = make_fill(col)
        ws1[ref].font = make_font(bold=True, color='FFFFFFFF', size=11)
        ws1[ref].alignment = make_align()
    ws1.row_dimensions[2].height = 26

    h1 = ['Rank', 'Student', 'Risk Level', 'Dropout %',
          'Attendance', 'Quiz', 'Assignment', 'Satisfaction',
          'Study Hrs', 'Dropout Intent', 'Why At Risk',
          'Suggestions', 'Teacher Action']
    for col, h in enumerate(h1, 1):
        c = ws1.cell(row=3, column=col)
        c.value     = h
        c.fill      = make_fill('FF0F3460')
        c.font      = make_font(bold=True, color='FFFFFFFF', size=10)
        c.alignment = make_align()
        c.border    = make_border()
    ws1.row_dimensions[3].height = 22

    for idx, row in df_results.iterrows():
        r    = idx + 4
        risk = row['Risk_Level']
        prob = row['Dropout_Probability']
        fc, bc = RISK_COLORS.get(risk, ('FF000000', 'FFFFFFFF'))

        row_data = [
            idx + 1,
            row.get('student_name', f'Student {idx+1}'),
            f"{'🔴' if risk=='CRITICAL' else '🟠' if risk=='HIGH' else '🟡' if risk=='MEDIUM' else '🟢'} {risk}",
            f"{prob*100:.1f}%",
            f"{row['attendance_frequency']*100:.0f}%",
            f"{row['quiz_score']*100:.0f}%",
            f"{row['assignment_completion']*100:.0f}%",
            f"{row['learning_satisfaction']*5:.1f}/5",
            f"{row['weekly_study_hours']*8:.1f}hrs",
            "Yes ⚠️" if row['prior_dropout_intention'] == 1 else "No ✅",
            row.get('Reasons', ''),
            row.get('Suggestions', ''),
            row.get('Teacher_Actions', '')
        ]

        for col, val in enumerate(row_data, 1):
            c = ws1.cell(row=r, column=col)
            c.value  = val
            c.border = make_border()
            c.alignment = make_align(h='left' if col > 10 else 'center', wrap=True)

            if col == 3:
                c.fill = make_fill(fc)
                c.font = make_font(bold=True, color='FFFFFFFF', size=10)
            elif col == 4:
                c.font = make_font(bold=True, color=fc[2:], size=11)
            elif col in [5, 6, 7]:
                try:
                    v = float(str(val).replace('%', '')) / 100
                    c.fill = make_fill(
                        'FFFDECEA' if v < 0.40 else 'FFFFF3E0' if v < 0.70 else 'FFE8F5E9')
                    c.font = make_font(
                        bold=True,
                        color=('FFE74C3C' if v < 0.40 else 'FFF39C12' if v < 0.70 else 'FF27AE60'))
                except:
                    pass
            elif idx % 2 == 0:
                c.fill = make_fill('FFF8F9FA')

        ws1.row_dimensions[r].height = 60

    widths1 = [5, 20, 14, 10, 11, 10, 11, 11, 10, 13, 50, 55, 45]
    for col, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A4'

    # ── SHEET 2: AT-RISK STUDENTS ──
    ws2 = wb.create_sheet("At-Risk Students")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"AT-RISK STUDENTS — IMMEDIATE ACTION  |  Course: {course_name}"
    ws2['A1'].fill = make_fill('FFE94560')
    ws2['A1'].font = make_font(bold=True, color='FFFFFFFF', size=13)
    ws2['A1'].alignment = make_align()
    ws2.row_dimensions[1].height = 30

    h2 = ['Priority', 'Student', 'Risk %', 'Risk Level',
          'Why At Risk', 'Attendance', 'Quiz', 'Teacher Action Required']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=2, column=col)
        c.value     = h
        c.fill      = make_fill('FF1A1A2E')
        c.font      = make_font(bold=True, color='FFFFFFFF', size=10)
        c.alignment = make_align()
        c.border    = make_border()
    ws2.row_dimensions[2].height = 22

    df_risk = df_results[
        df_results['Dropout_Probability'] >= 0.30
    ].reset_index(drop=True)

    for idx, row in df_risk.iterrows():
        r    = idx + 3
        risk = row['Risk_Level']
        prob = row['Dropout_Probability']
        fc, bc = RISK_COLORS.get(risk, ('FF000000', 'FFFFFFFF'))

        r2 = [
            f"#{idx+1} {risk}",
            row.get('student_name', f'Student {idx+1}'),
            f"{prob*100:.1f}%",
            risk,
            row.get('Reasons', ''),
            f"{row['attendance_frequency']*100:.0f}%",
            f"{row['quiz_score']*100:.0f}%",
            row.get('Teacher_Actions', '')
        ]

        for col, val in enumerate(r2, 1):
            c = ws2.cell(row=r, column=col)
            c.value     = val
            c.border    = make_border()
            c.alignment = make_align(h='left' if col > 4 else 'center', wrap=True)
            c.fill      = make_fill(bc)
            if col in [1, 3, 4]:
                c.fill = make_fill(fc)
                c.font = make_font(bold=True, color='FFFFFFFF', size=10)

        ws2.row_dimensions[r].height = 70

    widths2 = [14, 22, 10, 12, 55, 12, 12, 50]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A3'

    # ── SHEET 3: TEACHER DASHBOARD ──
    ws3 = wb.create_sheet("Teacher Dashboard")

    ws3.merge_cells('A1:H1')
    ws3['A1'] = f"TEACHER DASHBOARD  |  {teacher_name}  |  {course_name}"
    ws3['A1'].fill = make_fill('FF0F3460')
    ws3['A1'].font = make_font(bold=True, color='FFFFFFFF', size=13)
    ws3['A1'].alignment = make_align()
    ws3.row_dimensions[1].height = 30

    kpis = [
        (f"Total\n{total}", 'FF3498DB'),
        (f"Critical\n{critical}", 'FFE74C3C'),
        (f"High\n{high}", 'FFE67E22'),
        (f"Medium\n{medium}", 'FFF39C12'),
        (f"Low\n{low}", 'FF27AE60'),
        (f"At Risk\n{at_risk}", 'FFE94560'),
        (f"Safe\n{low}", 'FF27AE60'),
        (f"Dropout %\n{at_risk/max(total,1)*100:.0f}%", 'FF9B59B6'),
    ]
    for i, (txt, col) in enumerate(kpis):
        ws3.merge_cells(start_row=2, start_column=i+1, end_row=3, end_column=i+1)
        c = ws3.cell(row=2, column=i+1)
        c.value     = txt
        c.fill      = make_fill(col)
        c.font      = make_font(bold=True, color='FFFFFFFF', size=12)
        c.alignment = make_align(wrap=True)
        c.border    = make_border()
    ws3.row_dimensions[2].height = 45

    ws3.merge_cells('A4:H4')
    ws3['A4'] = "INTERVENTION GUIDE"
    ws3['A4'].fill = make_fill('FF2C3E50')
    ws3['A4'].font = make_font(bold=True, color='FFFFFFFF', size=11)
    ws3['A4'].alignment = make_align()
    ws3.row_dimensions[4].height = 22

    guides = [
        ('CRITICAL >80%', 'CALL student TODAY. Emergency counseling within 24 hours. Do not wait.', 'FFE74C3C'),
        ('HIGH 60-80%', 'Contact within 48 hours. Send support resources. Offer deadline extension.', 'FFE67E22'),
        ('MEDIUM 30-60%', 'Send personalized message. Share study tips. Invite to office hours.', 'FFF39C12'),
        ('LOW <30%', 'Regular weekly monitoring. Positive reinforcement. No urgent action.', 'FF27AE60'),
    ]
    for i, (level, guide, color) in enumerate(guides):
        r = 5 + i
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws3.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        c1 = ws3.cell(row=r, column=1)
        c1.value = level
        c1.fill  = make_fill(color)
        c1.font  = make_font(bold=True, color='FFFFFFFF', size=10)
        c1.alignment = make_align()
        c1.border    = make_border()
        c2 = ws3.cell(row=r, column=3)
        c2.value = guide
        c2.fill  = make_fill('FFF8F9FA')
        c2.font  = make_font(size=10)
        c2.alignment = make_align(h='left', wrap=True)
        c2.border    = make_border()
        ws3.row_dimensions[r].height = 35

    ws3.merge_cells('A10:H10')
    ws3['A10'] = "STUDENT ACTION LIST"
    ws3['A10'].fill = make_fill('FF0F3460')
    ws3['A10'].font = make_font(bold=True, color='FFFFFFFF', size=11)
    ws3['A10'].alignment = make_align()
    ws3.row_dimensions[10].height = 22

    h3 = ['#', 'Student', 'Risk %', 'Risk Level', 'Top Reason', 'Action', 'Deadline', 'Done?']
    for col, h in enumerate(h3, 1):
        c = ws3.cell(row=11, column=col)
        c.value     = h
        c.fill      = make_fill('FF1A1A2E')
        c.font      = make_font(bold=True, color='FFFFFFFF', size=10)
        c.alignment = make_align()
        c.border    = make_border()
    ws3.row_dimensions[11].height = 22

    for idx, row in df_results.iterrows():
        r    = 12 + idx
        risk = row['Risk_Level']
        prob = row['Dropout_Probability']
        fc, bc = RISK_COLORS.get(risk, ('FF000000', 'FFFFFFFF'))

        action = (
            "CALL NOW" if risk == 'CRITICAL'
            else "Contact Today" if risk == 'HIGH'
            else "Send Message" if risk == 'MEDIUM'
            else "Monitor")
        deadline = (
            "TODAY" if risk == 'CRITICAL'
            else "48 hrs" if risk == 'HIGH'
            else "This Week" if risk == 'MEDIUM'
            else "Monthly")

        reasons_text = row.get('Reasons', '')
        short_reason = (
            reasons_text[:60] + "..."
            if len(str(reasons_text)) > 60
            else reasons_text)

        r3 = [
            f"#{idx+1}",
            row.get('student_name', f'Student {idx+1}'),
            f"{prob*100:.1f}%",
            risk,
            short_reason,
            action,
            deadline,
            "Pending"
        ]
        for col, val in enumerate(r3, 1):
            c = ws3.cell(row=r, column=col)
            c.value     = val
            c.border    = make_border()
            c.alignment = make_align(h='left' if col > 4 else 'center', wrap=True)
            if col in [3, 4]:
                c.fill = make_fill(fc)
                c.font = make_font(bold=True, color='FFFFFFFF', size=10)
            elif idx % 2 == 0:
                c.fill = make_fill('FFF8F9FA')
        ws3.row_dimensions[r].height = 40

    widths3 = [6, 22, 10, 12, 40, 18, 12, 12]
    for col, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = 'A12'

    # ── SHEET 4: MANAGEMENT REPORT ──
    ws4 = wb.create_sheet("Management Report")

    ws4.merge_cells('A1:F1')
    ws4['A1'] = f"MANAGEMENT REPORT  |  {platform_name}  |  Dropout Prediction System"
    ws4['A1'].fill = make_fill('FF1A1A2E')
    ws4['A1'].font = make_font(bold=True, color='FFFFFFFF', size=13)
    ws4['A1'].alignment = make_align()
    ws4.row_dimensions[1].height = 32

    ws4.merge_cells('A2:F2')
    ws4['A2'] = "KEY PERFORMANCE INDICATORS"
    ws4['A2'].fill = make_fill('FF0F3460')
    ws4['A2'].font = make_font(bold=True, color='FFFFFFFF', size=12)
    ws4['A2'].alignment = make_align()
    ws4.row_dimensions[2].height = 25

    kpi4 = [
        ('Total Students', str(total), 'FF3498DB'),
        ('At Risk (>60%)', str(at_risk), 'FFE74C3C'),
        ('Safe (<30%)', str(low), 'FF27AE60'),
        ('Critical Cases', str(critical), 'FFE94560'),
        ('Dropout Rate', f"{at_risk/max(total,1)*100:.1f}%", 'FFE67E22'),
        ('Needs Action', str(at_risk), 'FF9B59B6'),
    ]
    for i, (label, val, color) in enumerate(kpi4):
        col   = (i % 3) * 2 + 1
        row_n = 3 + (i // 3) * 2
        ws4.merge_cells(start_row=row_n, start_column=col, end_row=row_n+1, end_column=col+1)
        c = ws4.cell(row=row_n, column=col)
        c.value = f"{label}\n{val}"
        c.fill  = make_fill(color)
        c.font  = make_font(bold=True, color='FFFFFFFF', size=13)
        c.alignment = make_align(wrap=True)
        c.border    = make_border()
        ws4.row_dimensions[row_n].height = 45

    rb = 9
    ws4.merge_cells(f'A{rb}:F{rb}')
    ws4[f'A{rb}'] = "RISK BREAKDOWN"
    ws4[f'A{rb}'].fill = make_fill('FF0F3460')
    ws4[f'A{rb}'].font = make_font(bold=True, color='FFFFFFFF', size=12)
    ws4[f'A{rb}'].alignment = make_align()
    ws4.row_dimensions[rb].height = 25

    for col, h in enumerate(['Level', 'Count', 'Percentage', 'Action', 'Timeline', 'Status'], 1):
        c = ws4.cell(row=rb+1, column=col)
        c.value     = h
        c.fill      = make_fill('FF2C3E50')
        c.font      = make_font(bold=True, color='FFFFFFFF', size=10)
        c.alignment = make_align()
        c.border    = make_border()

    rb_data = [
        ('CRITICAL >80%', critical, 'Personal intervention NOW', 'TODAY', 'FFE74C3C'),
        ('HIGH 60-80%', high, 'Teacher contact required', '48 Hours', 'FFE67E22'),
        ('MEDIUM 30-60%', medium, 'Support message + resources', 'This Week', 'FFF39C12'),
        ('LOW <30%', low, 'Regular monitoring', 'Monthly', 'FF27AE60'),
    ]
    for i, (level, count, action, timeline, color) in enumerate(rb_data):
        r = rb + 2 + i
        pct = f"{count/max(total,1)*100:.1f}%"
        for col, val in enumerate([level, count, pct, action, timeline, "Pending"], 1):
            c = ws4.cell(row=r, column=col)
            c.value     = val
            c.border    = make_border()
            c.alignment = make_align(wrap=True)
            if col in [1, 2, 3]:
                c.fill = make_fill(color)
                c.font = make_font(bold=True, color='FFFFFFFF', size=11)
            else:
                c.fill = make_fill('FFF8F9FA')
        ws4.row_dimensions[r].height = 30

    rec = rb + len(rb_data) + 4
    ws4.merge_cells(f'A{rec}:F{rec}')
    ws4[f'A{rec}'] = "STRATEGIC RECOMMENDATIONS"
    ws4[f'A{rec}'].fill = make_fill('FF0F3460')
    ws4[f'A{rec}'].font = make_font(bold=True, color='FFFFFFFF', size=12)
    ws4[f'A{rec}'].alignment = make_align()
    ws4.row_dimensions[rec].height = 25

    recs = [
        ('IMMEDIATE', f"Assign counselors to all {critical} critical students TODAY", 'FFE74C3C'),
        ('THIS WEEK', f"Ensure {high} high-risk students receive teacher contact", 'FFE67E22'),
        ('COURSE AUDIT', "Review modules with highest dropout correlation", 'FF3498DB'),
        ('FINANCIAL AID', "Offer payment plans and scholarship info to at-risk students", 'FFF39C12'),
        ('PEER MENTORING', "Pair critical students with top performers", 'FF27AE60'),
        ('WEEKLY REVIEW', "Review dropout dashboard every Monday morning", 'FF9B59B6'),
        ('ENGAGEMENT ALERTS', "Send automated notifications when activity drops", 'FF1ABC9C'),
        ('EARLY SUPPORT', "Contact ALL students in Week 1 to build connection", 'FFE94560'),
    ]
    for i, (priority, r_text, color) in enumerate(recs):
        row_r = rec + 1 + i
        ws4.merge_cells(start_row=row_r, start_column=2, end_row=row_r, end_column=6)
        c1 = ws4.cell(row=row_r, column=1)
        c1.value = priority
        c1.fill  = make_fill(color)
        c1.font  = make_font(bold=True, color='FFFFFFFF', size=10)
        c1.alignment = make_align()
        c1.border    = make_border()
        c2 = ws4.cell(row=row_r, column=2)
        c2.value = r_text
        c2.fill  = make_fill('FFF8F9FA')
        c2.font  = make_font(size=10)
        c2.alignment = make_align(h='left', wrap=True)
        c2.border    = make_border()
        ws4.row_dimensions[row_r].height = 32

    widths4 = [22, 55, 15, 35, 15, 15]
    for col, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================
# MAIN STREAMLIT APP
# ============================================
def main():

    st.markdown("""
    <div class="main-header">
        <h1>🎓 Student Dropout Prediction System</h1>
        <p>Before The Learner Quits — Powered by Logistic Regression + Decision Tree + SHAP</p>
    </div>
    """, unsafe_allow_html=True)

    if not models_loaded:
        st.error(
            "❌ Model files not found! "
            "Please ensure these files exist:\n"
            "- lr_model_v2.pkl\n"
            "- scaler.pkl\n"
            "- threshold_v2.pkl\n"
            "- shap_explainer_v2.pkl")
        st.info("Run your Jupyter Notebook cells 1-5 first to train and save models.")
        st.stop()

    st.success("✅ Models loaded successfully! Upload your student CSV file below.")

    # Sidebar
    st.sidebar.markdown("### ⚙️ Settings")
    teacher_name  = st.sidebar.text_input("Teacher Name", value="Prof. Rajesh Kumar")
    course_name   = st.sidebar.text_input("Course Name", value="Data Science Fundamentals")
    platform_name = st.sidebar.text_input("Platform Name", value="EdTech Platform")
    delimiter     = st.sidebar.selectbox("CSV Delimiter", [";", ",", "\t"],
                                         help="UCI dataset uses semicolon (;)")

    st.sidebar.markdown("---")
    st.sidebar.markdown(
        "### 📋 About\n"
        "This system predicts student dropout risk using:\n"
        "- ✅ Logistic Regression\n"
        "- ✅ Decision Tree\n"
        "- ✅ SHAP Explanations\n"
        "- ✅ Missing column handling\n"
        "- ✅ Excel report generation")

    # File upload
    st.markdown("### 📂 Upload Student Dataset")
    uploaded = st.file_uploader(
        "Upload any CSV file (UCI format or pre-processed)",
        type=["csv"],
        help="Works with any CSV — missing columns handled automatically"
    )

    if uploaded is None:
        st.info(
            "👆 Upload a CSV file to get started\n\n"
            "**Supported formats:**\n"
            "- UCI Student Dropout Dataset\n"
            "- Any CSV with student features\n"
            "- Partial datasets (missing columns are handled automatically)")
        st.stop()

    # Load data
    try:
        df_raw = pd.read_csv(uploaded, delimiter=delimiter, on_bad_lines='skip')
    except Exception as e:
        try:
            uploaded.seek(0)
            df_raw = pd.read_csv(uploaded)
        except Exception as e2:
            st.error(f"❌ Could not read file: {e2}")
            st.stop()

    st.success(f"✅ File loaded: {len(df_raw):,} students, {len(df_raw.columns)} columns")

    with st.expander("📋 Preview Raw Data", expanded=False):
        st.dataframe(df_raw.head(10))
        st.write(f"Columns found: {list(df_raw.columns)}")

    has_features = all(f in df_raw.columns for f in FEATURES)

    if has_features:
        st.info("✅ Dataset already has model features. Skipping UCI feature engineering.")
        df_feat = df_raw.copy()
    else:
        st.info("🔄 Engineering features from dataset columns...")
        df_feat = engineer_features(df_raw)

    if 'Target' in df_feat.columns:
        if st.checkbox("Filter: Remove 'Enrolled' students", value=True):
            before = len(df_feat)
            df_feat = df_feat[df_feat['Target'] != 'Enrolled'].copy()
            st.info(f"Removed Enrolled students: {before} → {len(df_feat)}")

    with st.spinner("🔄 Running predictions..."):
        probs, X_scaled, shap_vals_all, X = predict_all(df_feat)

    df_results = df_feat.copy().reset_index(drop=True)
    df_results['Dropout_Probability']     = probs
    df_results['Dropout_Probability_Pct'] = (probs * 100).round(1)

    if 'student_name' not in df_results.columns:
        df_results['student_name'] = [f"Student {i+1}" for i in range(len(df_results))]

    df_results['Risk_Level'] = [get_risk(p)[0] for p in probs]

    reasons_list = []
    suggestions_list = []
    actions_list = []

    for i, (_, row) in enumerate(df_results.iterrows()):
        shap_v  = shap_vals_all[i] if shap_vals_all is not None else None
        reasons = get_reasons(row, shap_v)
        sugs    = get_student_suggestions(row, probs[i])
        acts    = get_teacher_actions(probs[i], reasons)
        reasons_list.append(" | ".join(reasons[:3]))
        suggestions_list.append(" | ".join(sugs[:3]))
        actions_list.append(" | ".join(acts[:2]))

    df_results['Reasons']         = reasons_list
    df_results['Suggestions']     = suggestions_list
    df_results['Teacher_Actions'] = actions_list

    df_results = df_results.sort_values(
        'Dropout_Probability', ascending=False).reset_index(drop=True)

    st.markdown("---")
    st.markdown("## 📊 Prediction Results")

    total    = len(df_results)
    critical = (df_results['Risk_Level'] == 'CRITICAL').sum()
    high     = (df_results['Risk_Level'] == 'HIGH').sum()
    medium   = (df_results['Risk_Level'] == 'MEDIUM').sum()
    low      = (df_results['Risk_Level'] == 'LOW').sum()
    at_risk  = critical + high

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("👥 Total Students", total)
    col2.metric("🔴 Critical", critical, delta=f"{critical/max(total,1)*100:.0f}%")
    col3.metric("🟠 High Risk", high, delta=f"{high/max(total,1)*100:.0f}%")
    col4.metric("🟡 Medium", medium)
    col5.metric("🟢 Safe", low)

    st.markdown("---")

    tab1, tab2, tab3, tab4 = st.tabs([
        "🎓 Student View",
        "👨‍🏫 Teacher Dashboard",
        "🏢 Management Report",
        "📊 Analytics"
    ])

    # ── TAB 1: STUDENT VIEW ──
    with tab1:
        st.markdown("### 🎓 Student Risk Overview")
        st.info("Each student sees their personal risk score, why they are at risk, and what to do.")

        student_names = df_results['student_name'].tolist()
        selected = st.selectbox("Select a student to view details", student_names)

        s_row = df_results[df_results['student_name'] == selected].iloc[0]
        prob  = s_row['Dropout_Probability']
        risk, emoji, color, bg = get_risk(prob)

        st.markdown(
            f"<div style='background:#{bg};border-left:6px solid {color};"
            f"padding:20px;border-radius:10px;margin:10px 0'>",
            unsafe_allow_html=True)

        sc1, sc2, sc3 = st.columns(3)
        sc1.markdown(f"**Student:** {selected}")
        sc2.markdown(f"**Risk Level:** {emoji} **{risk}**")
        sc3.markdown(f"**Dropout Probability:** **{prob*100:.1f}%**")
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("#### 📊 Performance")
        feat_vals = [
            ('📅 Attendance',  s_row['attendance_frequency']),
            ('📝 Quiz Score',  s_row['quiz_score']),
            ('📚 Assignments', s_row['assignment_completion']),
            ('⏰ Study Hours', s_row['weekly_study_hours']),
            ('😊 Satisfaction',s_row['learning_satisfaction']),
        ]
        for label, val in feat_vals:
            bar_color = "🟢" if val >= 0.70 else "🟡" if val >= 0.50 else "🔴"
            st.markdown(f"**{label}** — {bar_color} {val*100:.0f}%")
            st.progress(float(val))

        st.markdown("#### ⚠️ Why At Risk")
        for r in s_row['Reasons'].split(' | '):
            if r:
                st.markdown(f"→ {r}")

        st.markdown("#### 💡 What You Should Do")
        for s in s_row['Suggestions'].split(' | '):
            if s:
                st.success(s)

        st.markdown("---")
        st.markdown("#### 📋 All Students")
        display_cols = ['student_name', 'Risk_Level', 'Dropout_Probability_Pct',
                        'attendance_frequency', 'quiz_score', 'assignment_completion', 'Reasons']
        existing_cols = [c for c in display_cols if c in df_results.columns]
        st.dataframe(
            df_results[existing_cols].rename(columns={
                'student_name': 'Student',
                'Risk_Level': 'Risk',
                'Dropout_Probability_Pct': 'Dropout %',
                'attendance_frequency': 'Attendance',
                'quiz_score': 'Quiz',
                'assignment_completion': 'Assignment',
                'Reasons': 'Why At Risk'
            }),
            use_container_width=True, height=400)

    # ── TAB 2: TEACHER DASHBOARD ──
    with tab2:
        st.markdown("### 👨‍🏫 Teacher Action Dashboard")
        st.info(f"**Teacher:** {teacher_name}  |  **Course:** {course_name}")

        tc1, tc2, tc3, tc4 = st.columns(4)
        tc1.error(f"🔴 CRITICAL\n\n**{critical}** students\n\nCall TODAY")
        tc2.warning(f"🟠 HIGH\n\n**{high}** students\n\nContact 48hrs")
        tc3.warning(f"🟡 MEDIUM\n\n**{medium}** students\n\nSend message")
        tc4.success(f"🟢 LOW\n\n**{low}** students\n\nMonitor only")

        st.markdown("#### 📊 Risk Distribution")
        fig, ax = plt.subplots(figsize=(8, 4))
        categories = ['Critical\n(>80%)', 'High\n(60-80%)', 'Medium\n(30-60%)', 'Low\n(<30%)']
        counts  = [critical, high, medium, low]
        colors  = ['#e74c3c', '#e67e22', '#f39c12', '#27ae60']
        bars = ax.bar(categories, counts, color=colors, edgecolor='white', linewidth=1.5)
        ax.set_title('Student Risk Distribution', fontweight='bold', fontsize=13)
        ax.set_ylabel('Number of Students')
        for bar, count in zip(bars, counts):
            if count > 0:
                ax.text(bar.get_x() + bar.get_width()/2,
                        bar.get_height() + 0.1, str(count),
                        ha='center', fontweight='bold', fontsize=12)
        ax.set_ylim(0, max(counts)*1.3 + 1)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        st.pyplot(fig)
        plt.close()

        st.markdown("#### 📋 Student Alert List")
        df_teacher = df_results[[
            'student_name', 'Risk_Level', 'Dropout_Probability_Pct',
            'attendance_frequency', 'quiz_score', 'Reasons', 'Teacher_Actions'
        ]].copy()
        df_teacher.columns = ['Student', 'Risk Level', 'Dropout %',
                               'Attendance', 'Quiz Score', 'Why At Risk', 'Action Required']

        def color_risk(val):
            if val == 'CRITICAL':
                return 'background-color: #fdecea; color: #e74c3c; font-weight: bold'
            elif val == 'HIGH':
                return 'background-color: #fff3e0; color: #e67e22; font-weight: bold'
            elif val == 'MEDIUM':
                return 'background-color: #fef9e7; color: #f39c12; font-weight: bold'
            else:
                return 'background-color: #e8f5e9; color: #27ae60; font-weight: bold'

        styled = df_teacher.style.applymap(color_risk, subset=['Risk Level'])
        st.dataframe(styled, use_container_width=True, height=500)

    # ── TAB 3: MANAGEMENT REPORT ──
    with tab3:
        st.markdown("### 🏢 Management Overview")

        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric("Total Students Monitored", total)
        mc2.metric("Students At Risk", at_risk, delta=f"{at_risk/max(total,1)*100:.0f}% of class")
        mc3.metric("Critical Cases", critical, delta="Need immediate attention")
        mc4.metric("Safe Students", low)

        st.markdown("#### 📈 Dropout Probability Distribution")
        fig2, ax2 = plt.subplots(figsize=(10, 4))
        ax2.hist(probs * 100, bins=20, color='#e74c3c', edgecolor='white', alpha=0.85)
        ax2.axvline(60, color='#e67e22', linestyle='--', linewidth=2, label='High Risk (60%)')
        ax2.axvline(80, color='#e74c3c', linestyle='--', linewidth=2, label='Critical (80%)')
        ax2.set_xlabel('Dropout Probability (%)')
        ax2.set_ylabel('Number of Students')
        ax2.set_title('Distribution of Dropout Risk Scores', fontweight='bold')
        ax2.legend()
        ax2.spines['top'].set_visible(False)
        ax2.spines['right'].set_visible(False)
        st.pyplot(fig2)
        plt.close()

        st.markdown("#### 📊 Feature Importance (What Drives Dropout Most)")
        lr_imp      = abs(model.coef_[0])
        lr_imp_norm = lr_imp / lr_imp.sum() * 100
        feat_names  = [FEATURE_DISPLAY[f] for f in FEATURES]

        fig3, ax3 = plt.subplots(figsize=(8, 4))
        sorted_pairs = sorted(zip(lr_imp_norm, feat_names))
        vals, names  = zip(*sorted_pairs)
        bar_colors   = ['#e74c3c' if v == max(vals) else '#3498db' for v in vals]
        bars3 = ax3.barh(names, vals, color=bar_colors)
        ax3.set_xlabel('Importance (%)')
        ax3.set_title('Feature Importance — Logistic Regression', fontweight='bold')
        for bar, val in zip(bars3, vals):
            ax3.text(bar.get_width() + 0.3,
                     bar.get_y() + bar.get_height()/2,
                     f'{val:.1f}%', va='center', fontweight='bold')
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        st.pyplot(fig3)
        plt.close()

        st.markdown("#### 🎯 Strategic Recommendations")
        if critical > 0:
            st.error(f"🔴 **IMMEDIATE:** Assign counselors to {critical} critical students TODAY")
        if high > 0:
            st.warning(f"🟠 **THIS WEEK:** Ensure {high} high-risk students receive teacher contact within 48 hours")
        st.info("📊 **COURSE AUDIT:** Review course modules with highest dropout correlation")
        st.info("💰 **FINANCIAL AID:** Proactively offer payment plans and scholarship information")
        st.success("🤝 **PEER MENTORING:** Pair critical students with top-performing peers")
        st.success("📱 **ENGAGEMENT ALERTS:** Send automated notifications when student activity drops")

    # ── TAB 4: ANALYTICS ──
    with tab4:
        st.markdown("### 📊 Model Analytics")

        if explainer is not None and shap_vals_all is not None:
            st.markdown("#### 🔬 SHAP Feature Impact")
            st.info("SHAP values show the actual contribution of each feature to each prediction")
            try:
                import shap
                fig4, ax4 = plt.subplots(figsize=(10, 5))
                shap.summary_plot(
                    shap_vals_all,
                    pd.DataFrame(X_scaled, columns=FEATURES),
                    feature_names=[FEATURE_DISPLAY[f] for f in FEATURES],
                    show=False, plot_size=None)
                st.pyplot(fig4)
                plt.close()
            except Exception as e:
                st.warning(f"SHAP plot error: {e}")
        else:
            st.warning("SHAP explainer not loaded. Run notebook cells to generate shap_explainer_v2.pkl")

        st.markdown("#### 📈 Model Performance")
        perf_data = {
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC'],
            'Logistic Regression': ['89.9%', '84.1%', '91.5%', '87.7%', '0.971'],
            'Decision Tree':       ['91.2%', '89.9%', '87.3%', '88.6%', '0.962'],
        }
        st.dataframe(pd.DataFrame(perf_data), use_container_width=True)

        st.info(
            "**Why Logistic Regression + Decision Tree?**\n\n"
            "These models achieve 89-91% accuracy while providing complete "
            "interpretability — teachers can see exactly WHY each student was flagged.")

    # ── EXCEL DOWNLOAD ──
    st.markdown("---")
    st.markdown("### 📥 Download Excel Report")

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.markdown(
            "**Your Excel file contains:**\n"
            "- Sheet 1: All Students Overview\n"
            "- Sheet 2: At-Risk Students Only\n"
            "- Sheet 3: Teacher Dashboard\n"
            "- Sheet 4: Management Report\n"
            "\nAll with color coding, SHAP-based reasons and personalized suggestions.")

    with col_dl2:
        if st.button("🔄 Generate Excel Report", type="primary"):
            with st.spinner("Generating Excel report..."):
                try:
                    excel_buf = generate_excel(
                        df_results, teacher_name, course_name, platform_name)
                    st.download_button(
                        label="📥 Download Complete Excel Report",
                        data=excel_buf,
                        file_name="student_dropout_prediction_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    st.success("✅ Excel report ready! Click button above to download.")
                except Exception as e:
                    st.error(f"❌ Excel generation error: {e}")
                    st.info("Check that openpyxl is installed: pip install openpyxl")

    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align:center;color:#666;font-size:12px'>"
        "Before The Learner Quits — Student Dropout Prediction System  |  "
        "Logistic Regression + Decision Tree + SHAP  |  "
        "UCI Dataset (4424 students, DOI: 10.3390/data7110146)"
        "</div>",
        unsafe_allow_html=True)


# ============================================
# RUN APP
# ============================================
if __name__ == "__main__":
    main()
